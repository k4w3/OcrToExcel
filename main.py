import cv2
import pytesseract
import re
from openpyxl import load_workbook

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

excel_file = "mac_list.xlsx"
wb = load_workbook(excel_file)
ws = wb.active

mac_pattern = r"(?:[0-9A-Fa-f]{2}[:\-]){5}[0-9A-Fa-f]{2}"

cap = cv2.VideoCapture(0)

print("MACアドレスを読み取ってください。")
print("MACアドレスをExcelに追加するには 'y'、スキップするには 'n' を押してください。")

pending_mac = None  # 確認待ちのMACアドレス格納用

while True:
    ret, frame = cap.read()
    if not ret:
        break

    height, width = frame.shape[:2]

    # box_width, box_height = 300, 50
    box_width, box_height = 250, 40
    x1 = width // 2 - box_width // 2
    y1 = height // 2 - box_height // 2
    x2 = x1 + box_width
    y2 = y1 + box_height

    cv2.rectangle(frame, (x1, y1), (x2, y2), (255, 0, 0), 2)

    # 画像処理
    roi = frame[y1:y2, x1:x2]
    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    contrast = cv2.convertScaleAbs(gray, alpha=1.5, beta=10)
    binary = cv2.threshold(contrast, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]

    # 処理中の画像の表示
    display_width = 500
    display_height = 80

    roi_disp = cv2.resize(roi, (display_width, display_height))
    gray_disp = cv2.cvtColor(cv2.resize(gray, (display_width, display_height)), cv2.COLOR_GRAY2BGR)
    contrast_disp = cv2.cvtColor(cv2.resize(contrast, (display_width, display_height)), cv2.COLOR_GRAY2BGR)
    binary_disp = cv2.cvtColor(cv2.resize(binary, (display_width, display_height)), cv2.COLOR_GRAY2BGR)

    top_row = cv2.hconcat([roi_disp, gray_disp])
    bottom_row = cv2.hconcat([contrast_disp, binary_disp])
    combined_view = cv2.vconcat([top_row, bottom_row])

    cv2.imshow("ROI > Gray > Contrast > Binary", combined_view)


    config = r'--oem 3 --psm 7'
    text = pytesseract.image_to_string(binary, config=config)

    # OCR結果を小さく表示
    cv2.putText(frame, text.strip(), (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

    if pending_mac is None:
        # MACアドレス抽出してまだ確認待ちがないときだけ処理
        macs = re.findall(mac_pattern, text)
        macs = [mac.upper() for mac in set(macs) if len(mac) == 17]

        if macs:
            mac = macs[0] # 1つめだけ確認用にセット
            existing_macs = [cell.value for cell in ws['A'] if cell.value]

            if mac in existing_macs:
                print(f"認識されたMAC: {mac}（重複しています。スキップします。）")
                cv2.putText(frame, f"{mac} is Duplicate", (10, height - 20),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
                cv2.imshow("MAC OCR Scanner", frame)
                cv2.waitKey(1000)
            else:
                pending_mac = mac # 確認待ちに設定
                print(f"認識されたMAC: {pending_mac}")
                print("追加する場合は 'y' を、スキップする場合は 'n' を押してください。")

    else:
        # 確認待ちMACアドレスを画面に表示
        cv2.putText(frame, f"Add to Excel? [y/n]: {pending_mac}", (10, height - 20),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)

    cv2.imshow("MAC OCR Scanner", frame)

    # キー入力待ち
    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):
        break
    elif pending_mac is not None:
        if key == ord('y'):
            ws.append([pending_mac])
            wb.save(excel_file)
            print(f"-> Excelに追加しました: {pending_mac}")
            pending_mac = None
        elif key == ord('n'):
            print(f"-> スキップしました: {pending_mac}")
            pending_mac = None

cap.release()
cv2.destroyAllWindows()
